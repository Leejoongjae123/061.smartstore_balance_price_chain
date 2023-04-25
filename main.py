import openpyxl
import pandas as pd
from pyautogui import size
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import subprocess
import shutil
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from bs4 import BeautifulSoup
import time
import datetime
import pyautogui
import pyperclip
import csv
import sys
import os
import math
import requests
import re
import random
import chromedriver_autoinstaller
from PyQt5.QtWidgets import QWidget, QApplication,QTreeView,QFileSystemModel,QVBoxLayout,QPushButton,QInputDialog,QLineEdit,QMainWindow,QMessageBox,QFileDialog,QTextEdit
from PyQt5.QtCore import QCoreApplication
from selenium.webdriver import ActionChains
from datetime import datetime,date,timedelta
import numpy
import datetime
from window import Ui_MainWindow
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import pybase64
import bcrypt
import http.client
import json
import pprint

def get_token(productNo, api_id,api_pw):
    time_now = datetime.datetime.now() - datetime.timedelta(seconds=3)
    time_now_stamp = math.ceil(datetime.datetime.timestamp(time_now) * 1000)
    # print(time_now)
    # print(time_now_stamp)

    clientId = api_id  # client id
    clientSecret = api_pw  # client pw
    # clientId=self.clientid
    # clientSecret=self.clientkey
    # timestamp = 1643961623299
    timestamp = time_now_stamp
    # 밑줄로 연결하여 password 생성
    password = clientId + "_" + str(timestamp)
    # bcrypt 해싱
    hashed = bcrypt.hashpw(password.encode('utf-8'), clientSecret.encode('utf-8'))
    # base64 인코딩
    result = pybase64.standard_b64encode(hashed).decode('utf-8')
    # print(result)
    params = {
        "client_id": clientId,
        "timestamp": time_now_stamp,
        "client_secret_sign": result,
        "grant_type": "client_credentials",
        "type": "SELF"
    }
    res = requests.post('https://api.commerce.naver.com/external/v1/oauth2/token', params=params)
    res.raise_for_status()

    token = eval(res.text)['access_token']

    conn = http.client.HTTPSConnection("api.commerce.naver.com")
    headers = {'Authorization': "Bearer {}".format(token)}
    conn.request("GET", "/external/v2/products/channel-products/{}".format(productNo), headers=headers)
    res = conn.getresponse()
    data = res.read()
    result = data.decode("utf-8")


    json_new_result = json.loads(result)
    json_new_result['originProduct']['detailContent']=""
    # pprint.pprint(json_new_result)
    json_new_result['originProduct']['statusType']="SALE"
    options=json_new_result['originProduct']['detailAttribute']['optionInfo']['optionCombinations']
    # pprint.pprint(options)

    for index_option,option in enumerate(options):
        origin_price=json_new_result['originProduct']['salePrice']
        origin_additional_price=option['price']
        try:
            managerCode=option['sellerManagerCode']
        except:
            print("매니저코드없음")
            continue
        if managerCode.find("/")>=0: #타겟의 옵션이 있는 경우
            managerCodeList=managerCode.split("/")
            origin_quantity=option['stockQuantity']
            print('managerCodeList:',managerCodeList,"본품기본가격:",origin_price,"본품옵션가격:",origin_additional_price,'본품재고:',origin_quantity)
            #타겟의 상품정보 가져옴
            try:
                conn = http.client.HTTPSConnection("api.commerce.naver.com")
                headers = {'Authorization': "Bearer {}".format(token)}
                conn.request("GET", "/external/v2/products/channel-products/{}".format(managerCodeList[0]), headers=headers)
                res = conn.getresponse()
                data = res.read()
                result = data.decode("utf-8")
                json_new_result2 = json.loads(result)
                #------------------------
                options_target = json_new_result2['originProduct']['detailAttribute']['optionInfo']['optionCombinations']
                # pprint.pprint(options_target)
            except:
                print("상품정보없음")
                continue
            for index_option_target,option_target in enumerate(options_target):
                managerCodeTarget=option_target['sellerManagerCode']
                # print('managerCodeTarget:',managerCodeTarget)
                if managerCodeTarget==managerCodeList[1]:
                    target_price=json_new_result2['originProduct']['salePrice']
                    target_option_price=option_target['price']
                    target_balance=option_target['stockQuantity']
                    print("타겟본가격:",target_price,"타겟옵션가격:",target_option_price,"타겟옵션재고:",target_balance)
                    # json_new_result['originProduct']['detailAttribute']['optionInfo']['optionCombinations'][index_option]=option_target['stockQuantity']
                    json_new_result['originProduct']['detailAttribute']['optionInfo']['optionCombinations'][index_option]['stockQuantity']=target_balance
                    json_new_result['originProduct']['detailAttribute']['optionInfo']['optionCombinations'][index_option]['price']=target_price+target_option_price-origin_price
        else: #타겟의 옵션이 없는경우
            origin_quantity = option['stockQuantity']
            print('managerCodeList:', managerCode, "본품기본가격:", origin_price, "본품옵션가격:", origin_additional_price, '본품재고:',origin_quantity)

            #타겟의 상품정보 가져옴
            conn = http.client.HTTPSConnection("api.commerce.naver.com")
            headers = {'Authorization': "Bearer {}".format(token)}
            conn.request("GET", "/external/v2/products/channel-products/{}".format(managerCode), headers=headers)
            res = conn.getresponse()
            data = res.read()
            result = data.decode("utf-8")
            json_new_result2 = json.loads(result)
            #---------------------
            target_balance=json_new_result2['originProduct']['stockQuantity']
            target_price=json_new_result2['originProduct']['salePrice']
            print('target_price:',target_price,'target_balance:',target_balance)
            json_new_result['originProduct']['detailAttribute']['optionInfo']['optionCombinations'][index_option]['stockQuantity']=target_balance
            json_new_result['originProduct']['detailAttribute']['optionInfo']['optionCombinations'][index_option]['price']=target_price-origin_price

        time.sleep(0.5)



    file_path = 'result.json'
    with open(file_path, 'w') as f:
        json.dump(json_new_result, f)

    token_path = 'token.txt'
    f = open(token_path, 'w')
    f.write(token)
    f.close()
    print("겟토큰완료")
def change_price(productNo):
    token_path = 'token.txt'
    with open(token_path) as f:
        lines = f.readlines()
        token = lines[0].strip()

    file_path = 'result.json'
    with open(file_path, 'r') as f:
        data = json.load(f)

    headers = {
        'Authorization': token,
        'content-type': "application/json"
    }

    # pprint.pprint(data)
    print("PUT요청 보내기")
    res = requests.put(
        'https://api.commerce.naver.com/external/v2/products/channel-products/{}'.format(productNo),
        data=json.dumps(data), headers=headers)
    print("PUT요청 완료")
    # res.raise_for_status()
    result = res.status_code
    print('result:',result ,res.text)
def load_excel(fname):
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    no_row = ws.max_row
    print("행갯수:", no_row)
    data_list = []
    for i in range(2, no_row + 1):
        product_no = ws.cell(row=i, column=1).value
        if product_no == "" or product_no == None:
            print('데이타 더 이상 없음')
            break
        start_flag=ws.cell(row=i, column=3).value
        data = [product_no,start_flag]
        data_list.append(data)
    print(data_list)
    return data_list
def load_excel_login(fname):
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    api_id=ws.cell(row=2,column=5).value
    api_pw = ws.cell(row=2, column=6).value
    print('api_id:',api_id,"api_pw:",api_pw)
    return api_id,api_pw


class Thread(QThread):
    cnt = 0
    user_signal = pyqtSignal(str)  # 사용자 정의 시그널 2 생성
    user_signal2 = pyqtSignal()  # 사용자 정의 시그널 2 생성

    def __init__(self, parent,fname):  # parent는 WndowClass에서 전달하는 self이다.(WidnowClass의 인스턴스)
        super().__init__(parent)
        self.parent = parent  # self.parent를 사용하여 WindowClass 위젯을 제어할 수 있다.
        self.fname=fname

    def run(self):
        product_no_list=load_excel(self.fname)
        api_id,api_pw=load_excel_login(self.fname)
        for index,product_no_elem in enumerate(product_no_list):
            text="{}번째 상품 변경중...".format(index+1)
            productNo=product_no_elem[0]
            start_flag = product_no_elem[1]
            if str(start_flag)=="2":
                print("미변경상품으로 건너뜀:{}".format(productNo))
                continue
            self.user_signal.emit(text)
            get_token(productNo, api_id, api_pw)
            change_price(productNo)

        self.user_signal2.emit()
    def stop(self):
        pass

class Example(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.path = "C:"
        self.index = None
        self.setupUi(self)
        self.setSlot()
        self.show()
        QApplication.processEvents()
        self.fname=""

    def start(self):
        print('11')
        self.x = Thread(self,self.fname)
        self.x.user_signal.connect(self.slot1)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.user_signal2.connect(self.slot2)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.start()

    def slot1(self, data1):  # 사용자 정의 시그널1에 connect된 function
        self.textEdit.append(str(data1))

    def slot2(self):  # 사용자 정의 시그널1에 connect된 function
        QMessageBox.information(self, "완료창", "작업이 완료 되었습니다.")

    def find(self):
        print("find")
        self.fname=QFileDialog.getOpenFileName(self," Open file",' ./')[0]
        print(self.fname)
        self.lineEdit.setText(self.fname)

    def setSlot(self):
        pass

    def setIndex(self, index):
        pass

    def quit(self):
        QCoreApplication.instance().quit()


app = QApplication([])
ex = Example()
sys.exit(app.exec_())



